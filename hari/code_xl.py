# -*- coding: utf-8 -*-
"""
Created on Wed Sep  4 22:11:23 2019

@author: Thangarajan
"""

import openpyxl 
from openpyxl.styles import PatternFill
path = "hari.xlsx"

wb_obj = openpyxl.load_workbook(path) 

cond_col = 'C'

sheet = wb_obj["Sheet1"]

row_count = sheet.max_row
column_count = sheet.max_column

for rows in sheet.iter_rows(min_row=2, max_col=column_count, max_row=row_count):
#    print(rows.cell)
    for r in rows:
#        print(r.value)
        cell_id = '{}{}'.format(r.column,r.row)
#        print(cell_id)
        if r.column == 'C':
            if type(r.value) is int and r.value < 3000:
                print(r.value)
                sheet[cell_id].fill = PatternFill(start_color="FFC7CE", fill_type = "solid")
                
wb_obj.save('result.xlsx')